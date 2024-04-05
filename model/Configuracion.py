from openpyxl import load_workbook


class Configuracion:
    def __init__(self, archivo_parametros):
        # Inicialización de la clase Configuracion con la ruta del archivo de parámetros de Excel
        self.archivo_parametros = archivo_parametros
        self.ruta_pdfs = None
        self.ruta_salida = None
        self.parametros = []
        self.cargar_parametros()  # Carga los parámetros al instanciar la clase

    def cargar_parametros(self):
        # Carga los datos del archivo Excel en la configuración
        wb = load_workbook(self.archivo_parametros)
        ws = wb['Rutas']

        # Lee las rutas de los PDFs y la ruta de salida desde la hoja 'Rutas'
        self.ruta_pdfs = ws['B2'].value
        self.ruta_salida = ws['B3'].value

        ws = wb['Documentos']
        # Itera sobre cada fila de la hoja 'Documentos' para obtener los parámetros de configuración
        for fila in ws.iter_rows(min_row=2, values_only=True):
            # Almacena los parámetros en un diccionario y los agrega a la lista de parámetros
            self.parametros.append({
                'tipo_documento': fila[0],  # Tipo de documento a procesar
                'prefijo': fila[1],  # Prefijo para usar en el nombramiento del archivo
                'expresion_regular': fila[2],  # Expresión regular para identificar el patrón
                'texto_verificacion': fila[3]  # Texto adicional para verificación
            })
